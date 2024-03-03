import openpyxl


class HomePageData:

    test_HomePage_data=[{"firstname":"Rahul","email":"Shetty","gender":"Male"}, {"firstname":"Naresh","email":"Talesha","gender":"Male"}]

    @staticmethod
    def getTestData(test_case_name):
        #Dict = {}
        book = openpyxl.load_workbook("/Users/ntalesha/Desktop/TESTING.xlsx")
        sheet = book.active
        dict = {}
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == "TestCase2":
                for j in range(2, sheet.max_column + 1):  # to get columns
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dict]
